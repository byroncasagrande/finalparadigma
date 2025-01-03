import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import io

def generar_reporte(df, resultados_regresion=None, nombre_archivo="reporte_analisis.xlsx"):
    """
    Genera un reporte ejecutivo en Excel.
    """
    try:
        workbook = openpyxl.Workbook()
        hoja = workbook.active
        hoja.title = "Resumen del Análisis"

        # Estilos
        titulo_font = Font(bold=True, size=16)
        subtitulo_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        borde_delgado = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        alineacion_centrada = Alignment(horizontal='center')

        # Título del reporte
        hoja.cell(row=1, column=1).value = "Reporte de Análisis Exploratorio y Regresión"
        hoja.cell(row=1, column=1).font = titulo_font

        # Resumen del DataFrame
        hoja.cell(row=3, column=1).value = "Resumen del DataFrame"
        hoja.cell(row=3, column=1).font = subtitulo_font

        # Información del DataFrame
        hoja.cell(row=4, column=1).value = "Número de filas:"
        hoja.cell(row=4, column=2).value = df.shape[0]
        hoja.cell(row=5, column=1).value = "Número de columnas:"
        hoja.cell(row=5, column=2).value = df.shape[1]

        # Estadísticas descriptivas
        hoja.cell(row=7, column=1).value = "Estadísticas Descriptivas"
        hoja.cell(row=7, column=1).font = subtitulo_font
        
        # Escribe los encabezados de las estadisticas descriptivas.
        hoja.cell(row=8, column=1).value = "Variable"
        hoja.cell(row=8, column=2).value = "count"
        hoja.cell(row=8, column=3).value = "mean"
        hoja.cell(row=8, column=4).value = "std"
        hoja.cell(row=8, column=5).value = "min"
        hoja.cell(row=8, column=6).value = "25%"
        hoja.cell(row=8, column=7).value = "50%"
        hoja.cell(row=8, column=8).value = "75%"
        hoja.cell(row=8, column=9).value = "max"
        for col in range(1, 10):
            hoja.cell(row=8, column=col).font = header_font
            hoja.cell(row=8, column=col).border = borde_delgado
            hoja.cell(row=8, column=col).alignment = alineacion_centrada

        estadisticas = df.describe()
        fila = 9
        for columna in estadisticas.columns:
            hoja.cell(row=fila, column=1).value = columna
            hoja.cell(row=fila, column=1).border = borde_delgado
            for i, estadistico in enumerate(estadisticas.index):
                hoja.cell(row=fila, column=i+2).value = estadisticas.loc[estadistico, columna]
                hoja.cell(row=fila, column=i+2).border = borde_delgado
                hoja.cell(row=fila, column=i+2).alignment = alineacion_centrada

            fila += 1
        
        # Resultados de la regresión (si existen)
        if resultados_regresion:
            hoja.cell(row=hoja.max_row + 2, column=1).value = "Resultados de la Regresión"
            hoja.cell(row=hoja.max_row, column=1).font = subtitulo_font
            fila = hoja.max_row + 1
            for resultado in resultados_regresion:
                hoja.cell(row=fila, column=1).value = resultado
                fila += 1
        else:
            hoja.cell(row=hoja.max_row + 2, column=1).value = "No se realizaron regresiones en este análisis."
            hoja.cell(row=hoja.max_row, column=1).font = subtitulo_font

        # Guardar el libro en un buffer de bytes en memoria
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)  # Regresar al inicio del buffer

        return buffer  # Retornar el buffer de bytes
    except Exception as e:
        print(f"Error al generar el reporte: {e}")
        return None